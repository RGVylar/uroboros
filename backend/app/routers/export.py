"""Full data export as a single Excel workbook with one sheet per data type."""
import io
from datetime import date, datetime, time, timezone

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import select
from sqlalchemy.orm import Session, joinedload

from app.database import get_db
from app.deps import get_current_user
from app.models.body_measurement import BodyMeasurementLog
from app.models.creatine import CreatineLog
from app.models.diary import DiaryEntry
from app.models.exercise import ExerciseSession, ExerciseSessionEntry
from app.models.mood import MoodEntry
from app.models.supplement import SupplementLog, UserSupplement
from app.models.user import User
from app.models.water import WaterLog
from app.models.weight import WeightLog

router = APIRouter(prefix="/export", tags=["export"])

_LEVEL_LABEL = {1: "bajo", 2: "normal", 3: "bueno"}

# Header style: dark green background, white bold text
_HEADER_FILL = PatternFill("solid", fgColor="1A3A2A")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def _write_sheet(ws, headers: list[str], rows: list[list]) -> None:
    """Write headers + rows to a worksheet with basic styling."""
    ws.append(headers)
    header_row = ws[1]
    for cell in header_row:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    for row in rows:
        ws.append(row)
    # Auto-width (approximate)
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)


def _dt_filter(stmt, col, date_from, date_to):
    if date_from:
        stmt = stmt.where(col >= datetime.combine(date_from, time.min, tzinfo=timezone.utc))
    if date_to:
        stmt = stmt.where(col <= datetime.combine(date_to, time.max, tzinfo=timezone.utc))
    return stmt


def _date_filter(stmt, col, date_from, date_to):
    if date_from:
        stmt = stmt.where(col >= date_from)
    if date_to:
        stmt = stmt.where(col <= date_to)
    return stmt


@router.get("/full.xlsx")
def export_full_xlsx(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> StreamingResponse:
    """Export all user data as a single Excel file with one sheet per category."""
    uid = user.id
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # ── Diario ────────────────────────────────────────────────────────────────
    stmt = (
        select(DiaryEntry)
        .where(DiaryEntry.user_id == uid)
        .order_by(DiaryEntry.consumed_at)
    )
    stmt = _dt_filter(stmt, DiaryEntry.consumed_at, date_from, date_to)
    diary_rows = []
    for e in db.scalars(stmt).unique():
        dt = e.consumed_at
        diary_rows.append([
            dt.strftime("%Y-%m-%d"),
            dt.strftime("%H:%M"),
            e.meal_type.value if e.meal_type else "",
            e.product.name if e.product else f"#{e.product_id}",
            e.product.brand if e.product and e.product.brand else "",
            round(e.grams, 1),
            round(e.calories, 1),
            round(e.protein, 1),
            round(e.carbs, 1),
            round(e.fat, 1),
        ])
    ws = wb.create_sheet("Diario")
    _write_sheet(ws, ["Fecha", "Hora", "Comida", "Producto", "Marca", "Gramos", "Kcal", "Proteína (g)", "Carbos (g)", "Grasa (g)"], diary_rows)

    # ── Agua ──────────────────────────────────────────────────────────────────
    stmt = select(WaterLog).where(WaterLog.user_id == uid).order_by(WaterLog.logged_date)
    stmt = _date_filter(stmt, WaterLog.logged_date, date_from, date_to)
    agua_rows = [[w.logged_date.isoformat(), int(w.ml)] for w in db.scalars(stmt)]
    ws = wb.create_sheet("Agua")
    _write_sheet(ws, ["Fecha", "ml"], agua_rows)

    # ── Peso ──────────────────────────────────────────────────────────────────
    stmt = select(WeightLog).where(WeightLog.user_id == uid).order_by(WeightLog.logged_at)
    stmt = _dt_filter(stmt, WeightLog.logged_at, date_from, date_to)
    peso_rows = [
        [w.logged_at.strftime("%Y-%m-%d"), w.logged_at.strftime("%H:%M"), round(w.weight, 2)]
        for w in db.scalars(stmt)
    ]
    ws = wb.create_sheet("Peso")
    _write_sheet(ws, ["Fecha", "Hora", "Kg"], peso_rows)

    # ── Medidas ───────────────────────────────────────────────────────────────
    stmt = (
        select(BodyMeasurementLog)
        .where(BodyMeasurementLog.user_id == uid)
        .order_by(BodyMeasurementLog.logged_at)
    )
    stmt = _dt_filter(stmt, BodyMeasurementLog.logged_at, date_from, date_to)
    meas_list = list(db.scalars(stmt))
    all_keys: list[str] = []
    for m in meas_list:
        for k in (m.measurements or {}).keys():
            if k not in all_keys:
                all_keys.append(k)
    meas_rows = []
    for m in meas_list:
        row = [m.logged_at.strftime("%Y-%m-%d"), m.logged_at.strftime("%H:%M")]
        row += [m.measurements.get(k, "") for k in all_keys]
        meas_rows.append(row)
    ws = wb.create_sheet("Medidas")
    _write_sheet(ws, ["Fecha", "Hora"] + all_keys, meas_rows)

    # ── Ejercicio ─────────────────────────────────────────────────────────────
    stmt = (
        select(ExerciseSession)
        .where(ExerciseSession.user_id == uid)
        .options(
            joinedload(ExerciseSession.entries).joinedload(ExerciseSessionEntry.exercise)
        )
        .order_by(ExerciseSession.session_date)
    )
    stmt = _date_filter(stmt, ExerciseSession.session_date, date_from, date_to)
    ejercicio_rows = []
    for session in db.scalars(stmt).unique():
        for entry in session.entries:
            ejercicio_rows.append([
                session.session_date.isoformat(),
                entry.exercise.name if entry.exercise else f"#{entry.exercise_id}",
                round(entry.quantity, 1),
                entry.exercise.unit if entry.exercise else "",
                round(entry.calories, 1),
                round(session.total_calories, 1),
            ])
    ws = wb.create_sheet("Ejercicio")
    _write_sheet(ws, ["Fecha", "Ejercicio", "Cantidad", "Unidad", "Kcal quemadas", "Total día (kcal)"], ejercicio_rows)

    # ── Suplementos ───────────────────────────────────────────────────────────
    stmt = (
        select(SupplementLog.logged_date, UserSupplement.name)
        .join(UserSupplement, SupplementLog.supplement_id == UserSupplement.id)
        .where(SupplementLog.user_id == uid)
        .order_by(SupplementLog.logged_date, UserSupplement.name)
    )
    stmt = _date_filter(stmt, SupplementLog.logged_date, date_from, date_to)
    supp_rows = [[row.logged_date.isoformat(), row.name] for row in db.execute(stmt)]
    ws = wb.create_sheet("Suplementos")
    _write_sheet(ws, ["Fecha", "Suplemento"], supp_rows)

    # ── Creatina ──────────────────────────────────────────────────────────────
    stmt = (
        select(CreatineLog)
        .where(CreatineLog.user_id == uid)
        .order_by(CreatineLog.logged_date)
    )
    stmt = _date_filter(stmt, CreatineLog.logged_date, date_from, date_to)
    creatine_rows = [[c.logged_date.isoformat()] for c in db.scalars(stmt)]
    ws = wb.create_sheet("Creatina")
    _write_sheet(ws, ["Fecha"], creatine_rows)

    # ── Mood ──────────────────────────────────────────────────────────────────
    stmt = (
        select(MoodEntry)
        .where(MoodEntry.user_id == uid)
        .order_by(MoodEntry.entry_date)
    )
    stmt = _date_filter(stmt, MoodEntry.entry_date, date_from, date_to)
    mood_rows = [
        [
            m.entry_date.isoformat(),
            _LEVEL_LABEL.get(m.energy, "") if m.energy else "",
            _LEVEL_LABEL.get(m.digestion, "") if m.digestion else "",
            _LEVEL_LABEL.get(m.mood, "") if m.mood else "",
            m.notes or "",
        ]
        for m in db.scalars(stmt)
    ]
    ws = wb.create_sheet("Estado del día")
    _write_sheet(ws, ["Fecha", "Energía", "Digestión", "Ánimo", "Notas"], mood_rows)

    # ── Serialize ─────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    suffix = ""
    if date_from:
        suffix += f"_{date_from}"
    if date_to:
        suffix += f"_{date_to}"
    filename = f"uroboros_export{suffix}.xlsx"

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
